from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_
from datetime import datetime, timedelta, date
from typing import Optional, List
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from ..database import get_db
from .. import models, auth

router = APIRouter(prefix="/api/reports", tags=["Reports & Analytics"])

def get_report_data(db: Session, start_date: datetime, end_date: datetime):
    # Total Revenue (sum of total_amount in Sales)
    revenue = db.query(func.sum(models.Sale.total_amount)).filter(
        models.Sale.sale_date.between(start_date, end_date)
    ).scalar() or 0.0
    
    # Total Profit
    profit = db.query(func.sum(models.Sale.calculated_profit)).filter(
        models.Sale.sale_date.between(start_date, end_date)
    ).scalar() or 0.0
    
    # Total Purchases
    purchases = db.query(func.sum(models.Purchase.total_amount)).filter(
        models.Purchase.purchase_date.between(start_date, end_date)
    ).scalar() or 0.0
    
    # Inventory Value (Sum of product carton quantity * carton purchase price)
    inventory_val = db.query(
        func.sum(models.Product.current_cartons * models.Product.purchase_price_carton)
    ).filter(models.Product.is_deleted == False).scalar() or 0.0
    
    # Cash vs Credit Sales
    cash_sales = db.query(func.sum(models.Sale.total_amount)).filter(
        and_(
            models.Sale.sale_date.between(start_date, end_date),
            models.Sale.sale_type == "CASH"
        )
    ).scalar() or 0.0
    
    credit_sales = db.query(func.sum(models.Sale.total_amount)).filter(
        and_(
            models.Sale.sale_date.between(start_date, end_date),
            models.Sale.sale_type == "CREDIT"
        )
    ).scalar() or 0.0
    
    # Expenses (CashOutflows that are EXPENSE)
    expenses = db.query(func.sum(models.CashFlow.amount)).filter(
        and_(
            models.CashFlow.created_at.between(start_date, end_date),
            models.CashFlow.type == "OUTFLOW",
            models.CashFlow.source == "EXPENSE"
        )
    ).scalar() or 0.0
    
    net_profit = profit - expenses
    
    # Top Selling Products
    top_products = db.query(
        models.Product.name,
        func.sum(models.SaleItem.quantity_pieces).label("total_sold"),
        func.sum(models.SaleItem.quantity_pieces * models.SaleItem.price_per_piece).label("total_revenue")
    ).join(models.SaleItem, models.Product.id == models.SaleItem.product_id)\
     .join(models.Sale, models.Sale.id == models.SaleItem.sale_id)\
     .filter(models.Sale.sale_date.between(start_date, end_date))\
     .group_by(models.Product.name)\
     .order_by(func.sum(models.SaleItem.quantity_pieces).desc())\
     .limit(5).all()
     
    top_selling = [
        {"name": p[0], "total_sold": int(p[1]), "total_revenue": round(p[2], 2)}
        for p in top_products
    ]

    # Lowest Selling Products
    lowest_products = db.query(
        models.Product.name,
        func.sum(models.SaleItem.quantity_pieces).label("total_sold")
    ).join(models.SaleItem, models.Product.id == models.SaleItem.product_id)\
     .join(models.Sale, models.Sale.id == models.SaleItem.sale_id)\
     .filter(models.Sale.sale_date.between(start_date, end_date))\
     .group_by(models.Product.name)\
     .order_by(func.sum(models.SaleItem.quantity_pieces).asc())\
     .limit(5).all()
     
    lowest_selling = [
        {"name": p[0], "total_sold": int(p[1])}
        for p in lowest_products
    ]
    
    # Best Customers (Highest Purchases)
    best_cust = db.query(
        models.Customer.name,
        func.sum(models.Sale.total_amount).label("total_purchased")
    ).join(models.Sale, models.Customer.id == models.Sale.customer_id)\
     .filter(models.Sale.sale_date.between(start_date, end_date))\
     .group_by(models.Customer.name)\
     .order_by(func.sum(models.Sale.total_amount).desc())\
     .limit(5).all()
     
    best_customers = [
        {"name": c[0], "total_purchased": round(c[1], 2)}
        for c in best_cust
    ]
    
    # Highest Debt Customers
    debt_cust = db.query(
        models.Customer.name,
        models.Customer.phone,
        models.Customer.current_debt
    ).filter(
        models.Customer.is_deleted == False,
        models.Customer.current_debt > 0
    ).order_by(models.Customer.current_debt.desc())\
     .limit(5).all()
     
    highest_debt = [
        {"name": c[0], "phone": c[1], "debt": round(c[2], 2)}
        for c in debt_cust
    ]
    
    # Daily Sales Trend (last 30 days of selected period or full period if smaller)
    sales_trend_query = db.query(
        func.date(models.Sale.sale_date).label("date"),
        func.sum(models.Sale.total_amount).label("daily_revenue"),
        func.sum(models.Sale.calculated_profit).label("daily_profit")
    ).filter(models.Sale.sale_date.between(start_date, end_date))\
     .group_by(func.date(models.Sale.sale_date))\
     .order_by(func.date(models.Sale.sale_date).asc()).all()
     
    sales_trend = [
        {
            "date": str(row.date),
            "revenue": round(row.daily_revenue, 2),
            "profit": round(row.daily_profit, 2)
        } for row in sales_trend_query
    ]
    
    # Category Distribution
    category_dist_query = db.query(
        models.Product.category,
        func.sum(models.SaleItem.quantity_pieces * models.SaleItem.price_per_piece).label("cat_revenue")
    ).join(models.SaleItem, models.Product.id == models.SaleItem.product_id)\
     .join(models.Sale, models.Sale.id == models.SaleItem.sale_id)\
     .filter(models.Sale.sale_date.between(start_date, end_date))\
     .group_by(models.Product.category).all()
     
    category_distribution = [
        {"category": row.category or "غير محدد", "revenue": round(row.cat_revenue, 2)}
        for row in category_dist_query
    ]
    
    return {
        "revenue": round(revenue, 2),
        "profit": round(profit, 2),
        "purchases": round(purchases, 2),
        "inventory_value": round(inventory_val, 2),
        "cash_sales": round(cash_sales, 2),
        "credit_sales": round(credit_sales, 2),
        "expenses": round(expenses, 2),
        "net_profit": round(net_profit, 2),
        "top_selling": top_selling,
        "lowest_selling": lowest_selling,
        "best_customers": best_customers,
        "highest_debt": highest_debt,
        "sales_trend": sales_trend,
        "category_distribution": category_distribution
    }

@router.get("/summary")
def get_report_summary(
    range_type: str = "last_90_days",
    custom_start: Optional[str] = None,
    custom_end: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    end_dt = datetime.now()
    start_dt = end_dt - timedelta(days=90)
    
    if range_type == "today":
        start_dt = datetime.combine(date.today(), datetime.min.time())
    elif range_type == "yesterday":
        start_dt = datetime.combine(date.today() - timedelta(days=1), datetime.min.time())
        end_dt = datetime.combine(date.today() - timedelta(days=1), datetime.max.time())
    elif range_type == "last_7_days":
        start_dt = end_dt - timedelta(days=7)
    elif range_type == "last_30_days":
        start_dt = end_dt - timedelta(days=30)
    elif range_type == "this_month":
        start_dt = datetime(end_dt.year, end_dt.month, 1)
    elif range_type == "last_month":
        first_day_current = datetime(end_dt.year, end_dt.month, 1)
        end_dt = first_day_current - timedelta(seconds=1)
        start_dt = datetime(end_dt.year, end_dt.month, 1)
    elif range_type == "this_year":
        start_dt = datetime(end_dt.year, 1, 1)
    elif range_type == "custom" and custom_start and custom_end:
        try:
            start_dt = datetime.strptime(custom_start, "%Y-%m-%d")
            end_dt = datetime.strptime(custom_end, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
        except ValueError:
            raise HTTPException(status_code=400, detail="تنسيق تاريخ غير صحيح. يجب أن يكون YYYY-MM-DD")
            
    return get_report_data(db, start_dt, end_dt)

@router.get("/export/excel")
def export_excel(
    range_type: str = "last_90_days",
    custom_start: Optional[str] = None,
    custom_end: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    # Compute range dates
    end_dt = datetime.now()
    start_dt = end_dt - timedelta(days=90)
    
    if range_type == "today":
        start_dt = datetime.combine(date.today(), datetime.min.time())
    elif range_type == "last_7_days":
        start_dt = end_dt - timedelta(days=7)
    elif range_type == "last_30_days":
        start_dt = end_dt - timedelta(days=30)
    elif range_type == "custom" and custom_start and custom_end:
        start_dt = datetime.strptime(custom_start, "%Y-%m-%d")
        end_dt = datetime.strptime(custom_end, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
        
    data = get_report_data(db, start_dt, end_dt)
    
    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "التقرير المالي"
    ws.views.sheetView[0].showGridLines = True
    
    # RTL direction
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.sheet_view.rightToLeft = True
    
    # Styles
    title_font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")
    bold_font = Font(name="Segoe UI", size=11, bold=True)
    regular_font = Font(name="Segoe UI", size=11)
    
    title_fill = PatternFill(start_color="18181B", end_color="18181B", fill_type="solid")
    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    accent_fill = PatternFill(start_color="F4F4F5", end_color="F4F4F5", fill_type="solid")
    
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    border_side = Side(border_style="thin", color="D4D4D8")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Title Block
    ws.merge_cells("A1:D2")
    title_cell = ws["A1"]
    title_cell.value = "تقرير أداء السوبر ماركت الذكي"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_align
    
    ws["A3"] = "الفترة:"
    ws["A3"].font = bold_font
    ws["B3"] = f"من {start_dt.strftime('%Y-%m-%d')} إلى {end_dt.strftime('%Y-%m-%d')}"
    ws["B3"].font = regular_font
    
    # KPI Headers
    ws["A5"] = "المؤشر المالي"
    ws["A5"].font = header_font
    ws["A5"].fill = header_fill
    ws["A5"].alignment = center_align
    
    ws["B5"] = "القيمة (جنيه)"
    ws["B5"].font = header_font
    ws["B5"].fill = header_fill
    ws["B5"].alignment = center_align
    
    # Fill KPIs
    kpis = [
        ("إجمالي الإيرادات", data["revenue"]),
        ("إجمالي الأرباح التجارية", data["profit"]),
        ("إجمالي المشتريات", data["purchases"]),
        ("إجمالي المصاريف", data["expenses"]),
        ("صافي الربح", data["net_profit"]),
        ("قيمة المخزون الحالية", data["inventory_value"]),
    ]
    
    row_idx = 6
    for kpi, val in kpis:
        ws.cell(row=row_idx, column=1, value=kpi).font = bold_font if "صافي" in kpi else regular_font
        ws.cell(row=row_idx, column=1).alignment = right_align
        ws.cell(row=row_idx, column=1).border = thin_border
        
        val_cell = ws.cell(row=row_idx, column=2, value=val)
        val_cell.font = bold_font if "صافي" in kpi else regular_font
        val_cell.alignment = center_align
        val_cell.border = thin_border
        val_cell.number_format = '#,##0.00'
        
        if "صافي" in kpi:
            ws.cell(row=row_idx, column=1).fill = accent_fill
            val_cell.fill = accent_fill
            
        row_idx += 1
        
    # Top Products Section
    row_idx += 2
    ws.cell(row=row_idx, column=1, value="المنتجات الأكثر مبيعاً").font = header_font
    ws.cell(row=row_idx, column=1).fill = header_fill
    ws.cell(row=row_idx, column=1).alignment = center_align
    
    ws.cell(row=row_idx, column=2, value="الكمية المباعة (حبة)").font = header_font
    ws.cell(row=row_idx, column=2).fill = header_fill
    ws.cell(row=row_idx, column=2).alignment = center_align
    
    ws.cell(row=row_idx, column=3, value="إجمالي مبيعات المنتج").font = header_font
    ws.cell(row=row_idx, column=3).fill = header_fill
    ws.cell(row=row_idx, column=3).alignment = center_align
    
    row_idx += 1
    for prod in data["top_selling"]:
        c1 = ws.cell(row=row_idx, column=1, value=prod["name"])
        c1.font = regular_font
        c1.border = thin_border
        c1.alignment = right_align
        
        c2 = ws.cell(row=row_idx, column=2, value=prod["total_sold"])
        c2.font = regular_font
        c2.border = thin_border
        c2.alignment = center_align
        
        c3 = ws.cell(row=row_idx, column=3, value=prod["total_revenue"])
        c3.font = regular_font
        c3.border = thin_border
        c3.alignment = center_align
        c3.number_format = '#,##0.00'
        
        row_idx += 1
        
    # Save Workbook
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    filename = f"report_{range_type}_{start_dt.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
